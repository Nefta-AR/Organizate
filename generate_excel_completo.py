import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# HOJA 1: Resumen del Proyecto
# ============================================================
ws1 = wb.active
ws1.title = "Resumen del Proyecto"

# Informacion General
ws1['A1'] = "REPORTE DE PROYECTO: SIMPLE"
ws1['A3'] = "INFORMACION GENERAL"
ws1['A4'] = "Repositorio"
ws1['B4'] = "Nefta-AR/Organizate"
ws1['A5'] = "Plataforma"
ws1['B5'] = "Flutter + Firebase"
ws1['A6'] = "Periodo"
ws1['B6'] = "27 Abril 2026 - Julio 2026"
ws1['A7'] = "Estado General"
ws1['B7'] = "En Desarrollo (97% Completado)"

ws1['A9'] = "PROGRESO GENERAL: 97%"

# Fases del Proyecto
ws1['A11'] = "FASES DEL PROYECTO"
ws1['A12'] = "Fase"
ws1['B12'] = "Estado"
ws1['C12'] = "Progreso"

fases = [
    ("Fase 1: Fundacion y Auth", "Completado", "100%"),
    ("Fase 2: IA / Super Experto", "Completado", "100%"),
    ("Fase 3: Modulo TEA (Pictogramas)", "Completado", "100%"),
    ("Fase 4: Modulo TDAH (Tareas)", "Completado", "100%"),
    ("Fase 5: Integracion y Correcciones", "Completado", "100%"),
    ("Fase 6: Pulido y Testing", "En Curso", "75%"),
    ("Fase 7: Documentacion y Entrega", "Pendiente", "0%"),
]

for i, (fase, estado, progreso) in enumerate(fases, 13):
    ws1[f'A{i}'] = fase
    ws1[f'B{i}'] = estado
    ws1[f'C{i}'] = progreso

# Distribucion por Modulo
ws1['A20'] = "DISTRIBUCION DE TRABAJO POR MODULO"
ws1['A21'] = "Modulo"
ws1['B21'] = "Porcentaje (%)"
ws1['C21'] = "Estado"

modulos = [
    ("Autenticacion y base", 15, "Completado"),
    ("Modulo TEA (Pictogramas)", 30, "Completado"),
    ("Modulo TDAH (Tareas + Foco)", 20, "Completado"),
    ("IA / Super Experto", 8, "Completado"),
    ("Panel Tutor (supervision)", 12, "Completado"),
    ("Infraestructura Firebase", 5, "Completado"),
    ("Pulido y Dashboard", 10, "En Curso"),
]

for i, (modulo, pct, estado) in enumerate(modulos, 22):
    ws1[f'A{i}'] = modulo
    ws1[f'B{i}'] = pct
    ws1[f'C{i}'] = estado

# Metricas
ws1['A30'] = "METRICAS DEL PROYECTO"
ws1['A31'] = "Lineas de codigo"
ws1['B31'] = "~18,000+"
ws1['A32'] = "Archivos Dart"
ws1['B32'] = "65+"
ws1['A33'] = "Commits"
ws1['B33'] = "30+"
ws1['A34'] = "Funcionalidades"
ws1['B34'] = "50+"

# Column widths
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 15

# ============================================================
# HOJA 2: Carta Gantt
# ============================================================
ws2 = wb.create_sheet("Carta Gantt")

ws2['A1'] = "CARTA GANTT - CRONOGRAMA COMPLETO (27 Abril - 07 Julio 2026)"

# Headers
ws2['A3'] = "FASE"
ws2['B3'] = "TAREA"
ws2['C3'] = "ESTADO"

# Weekly columns headers (S1 to S12)
semanas = [
    ("D", "S1 27-30Abr"),
    ("E", "S2 01-04May"),
    ("F", "S3 05-09May"),
    ("G", "S4 10-13May"),
    ("H", "S5 14-18May"),
    ("I", "S6 19-23May"),
    ("J", "S7 24-30May"),
    ("K", "S8 31May-06Jun"),
    ("L", "S9 07-13Jun"),
    ("M", "S10 14-20Jun"),
    ("N", "S11 21-27Jun"),
    ("O", "S12 28Jun-07Jul"),
]

for col, label in semanas:
    ws2[f'{col}3'] = label

# Tasks data: (fase, tarea, estado, semanas_activas)
# semanas_activas is a list of column letters where the task is active
tareas = [
    # Fase 1
    ("FASE 1: Fundacion y Auth", "Base Organizate 2.0 + rediseno UI", "Completado", ["D"]),
    ("", "Sistema Login (email + Google)", "Completado", ["D"]),
    ("", "Cambio nombre -> Simple + Logo", "Completado", ["D"]),
    # Fase 2
    ("FASE 2: IA / Super Experto", "Integracion Cloud Functions IA", "Completado", ["E"]),
    ("", "Super Experto funcional (Gemini)", "Completado", ["E"]),
    # Fase 3
    ("FASE 3: Modulo TEA (Pictogramas)", "Pantalla Pictogramas Beta", "Completado", ["F"]),
    ("", "Banco pictogramas predefinidos (SVG)", "Completado", ["F"]),
    ("", "Pictogramas con color", "Completado", ["F"]),
    ("", "Crear pictogramas personalizados", "Completado", ["G"]),
    # Fase 4
    ("FASE 4: Modulo TDAH (Tareas)", "Gestion tareas (CRUD + categorias)", "Completado", ["F"]),
    ("", "Swipe para eliminar tareas", "Completado", ["F"]),
    ("", "Migracion completa a Simple", "Completado", ["F"]),
    ("", "Timer Pomodoro + respiracion", "Completado", ["G"]),
    ("", "Sistema puntos y racha", "Completado", ["G"]),
    # Fase 5
    ("FASE 5: Integracion", "Correccion SHA + Firebase", "Completado", ["G"]),
    ("", "Eliminacion Modo Foco", "Completado", ["G"]),
    ("", "Fix superposicion botones TEA", "Completado", ["G"]),
    ("", "Tutor conectado a paciente", "Completado", ["H"]),
    ("", "Supervision tutor -> detalle paciente", "Completado", ["H"]),
    ("", "Sincronizacion bidireccional", "Completado", ["I"]),
    ("", "Correccion bugs integracion", "Completado", ["I"]),
    # Fase 6
    ("FASE 6: Pulido", "Dashboard progreso (fl_chart)", "Completado", ["J"]),
    ("", "Control granular pestañas", "Completado", ["J"]),
    ("", "Fix nav bar feature flags", "Completado", ["J"]),
    ("", "Eliminacion pantalla legacy", "Completado", ["J"]),
    ("", "Notificaciones push FCM", "En Curso", ["K"]),
    ("", "Testing y bugs menores", "Pendiente", ["K"]),
    # Fase 7
    ("FASE 7: Documentacion", "Optimizacion rendimiento", "Pendiente", ["L"]),
    ("", "Documentacion tecnica", "Pendiente", ["L"]),
    ("", "Pruebas finales", "Pendiente", ["M"]),
    ("", "Manual de usuario", "Pendiente", ["M"]),
    ("", "Presentacion final", "Pendiente", ["N"]),
]

row = 4
for fase, tarea, estado, cols in tareas:
    ws2[f'A{row}'] = fase
    ws2[f'B{row}'] = tarea
    ws2[f'C{row}'] = estado
    for col in cols:
        ws2[f'{col}{row}'] = "X"
    row += 1

# Hitos section
row += 2
ws2[f'A{row}'] = "HITOS DEL PROYECTO"
row += 1
ws2[f'A{row}'] = "SEMANA"
ws2[f'B{row}'] = "HITO"
ws2[f'C{row}'] = "FECHA"
ws2[f'D{row}'] = "ESTADO"

hitos = [
    ("S3", "HITO 1: Vinculacion Tutor-Paciente", "13 May 2026", "ALCANZADO"),
    ("S5", "HITO 2: Sincronizacion Completa", "26 May 2026", "ALCANZADO"),
    ("S9", "HITO 3: MVP Listo", "30 Jun 2026", "PENDIENTE"),
    ("S12", "HITO 4: Entrega Final", "07 Jul 2026", "PENDIENTE"),
]

for h in hitos:
    row += 1
    ws2[f'A{row}'] = h[0]
    ws2[f'B{row}'] = h[1]
    ws2[f'C{row}'] = h[2]
    ws2[f'D{row}'] = h[3]

# Column widths
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 14
for col, _ in semanas:
    ws2.column_dimensions[col].width = 12

# ============================================================
# HOJA 3: Proximo Sprint
# ============================================================
ws3 = wb.create_sheet("Proximo Sprint")

ws3['A1'] = "PROXIMO SPRINT - TAREAS PENDIENTES"
ws3['A3'] = "Actualizado: 26 Mayo 2026"

ws3['A5'] = "Prioridad"
ws3['B5'] = "Tarea"
ws3['C5'] = "Fase"
ws3['D5'] = "Estado"

tareas_sprint = [
    ("Alta", "Notificaciones push FCM completas", "Fase 6", "En Curso"),
    ("Alta", "QA y testing en dispositivos reales", "Fase 6", "Pendiente"),
    ("Media", "Notificacion FCM al tutor cuando usuario completa tarea", "Fase 6", "Pendiente"),
    ("Media", "Documentacion tecnica completa", "Fase 7", "Pendiente"),
    ("Media", "Manual de usuario", "Fase 7", "Pendiente"),
    ("Media", "Presentacion final del proyecto", "Fase 7", "Pendiente"),
    ("Baja", "Optimizacion lazy loading imagenes", "Fase 7", "Pendiente"),
    ("Baja", "APK firmado y empaquetado", "Fase 7", "Pendiente"),
    ("Baja", "Video demo del proyecto", "Fase 7", "Pendiente"),
]

for i, (prio, tarea, fase, estado) in enumerate(tareas_sprint, 6):
    ws3[f'A{i}'] = prio
    ws3[f'B{i}'] = tarea
    ws3[f'C{i}'] = fase
    ws3[f'D{i}'] = estado

# Contexto de prioridades
row = 16
ws3[f'A{row}'] = "CONTEXTO DE PRIORIDADES:"
row += 1
ws3[f'A{row}'] = "Alta"
ws3[f'B{row}'] = "Critico para el funcionamiento"
row += 1
ws3[f'A{row}'] = "Media"
ws3[f'B{row}'] = "Mejora la experiencia de usuario"
row += 1
ws3[f'A{row}'] = "Baja"
ws3[f'B{row}'] = "Puede esperar al final"

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 14

# ============================================================
# HOJA 4: Hitos (Commits)
# ============================================================
ws4 = wb.create_sheet("Hitos (Commits)")

ws4['A1'] = "HISTORIAL DE HITOS (COMMITS)"

ws4['A3'] = "Fecha"
ws4['B3'] = "Commit"
ws4['C3'] = "Descripcion"
ws4['D3'] = "Fase"

commits = [
    ("27 Abr 2026", "8e73da2", "Base Organizate 2.0 + rediseno UI Login", "Fase 1"),
    ("28 Abr 2026", "e48e5d0", "Cambio de nombre a Simple, nuevo logo, login web", "Fase 1"),
    ("29 Abr 2026", "2d886cc", "Super Experto IA funcional (Gemini + Cloud Functions)", "Fase 2"),
    ("30 Abr 2026", "0633e86", "Pantalla Pictogramas Beta (modulo TEA)", "Fase 3"),
    ("05 May 2026", "12bd20d", "Migracion completa de Organizate -> Simple", "Fase 4"),
    ("06 May 2026", "335f814", "Pantalla Pictogramas completa", "Fase 3"),
    ("09 May 2026", "97fd1b1", "Eliminacion Modo Foco -> reemplazado por Pictogramas TEA", "Fase 5"),
    ("09 May 2026", "89ec05e", "Conexiones Firebase, nuevo SHA-1, google-services.json", "Fase 5"),
    ("13 May 2026", "ce5c88a", "Tutor conectado a paciente (vinculacion completada)", "Fase 5"),
    ("14 May 2026", "e6d1494", "Inicio de sesion con Google corregido", "Fase 5"),
    ("19 May 2026", "-", "ProfileSetupScreen post-rol + renombrado paciente -> usuario", "Fase 5"),
    ("19 May 2026", "-", "Panel historial tutor: stats + log Pomodoro + log pictogramas", "Fase 5"),
    ("19 May 2026", "-", "Badge Tutor en tareas + sincronizacion bidireccional", "Fase 5"),
    ("19 May 2026", "-", "CustomNavBar reactivo + tabs dinamicos", "Fase 6"),
    ("19 May 2026", "-", "SettingsScreen redisenado + fix cambio de rol", "Fase 6"),
    ("24 May 2026", "-", "Fix flujo auth completo (3 partes) + limpieza roles", "Fase 6"),
    ("24 May 2026", "-", "PantallasConfigScreen + defaults pestañas", "Fase 6"),
    ("24 May 2026", "-", "Ajustes tutor simplificado + contacto emergencia", "Fase 6"),
    ("26 May 2026", "-", "Dashboard progreso integrado (3 graficos)", "Fase 6"),
    ("26 May 2026", "-", "Fix CustomNavBar featureInicio/featureTareas", "Fase 6"),
    ("26 May 2026", "-", "Eliminada TutorPatientDetailScreen (legacy)", "Fase 6"),
    ("26 May 2026", "-", "Kiosk Mode removido del MVP (escalabilidad futura)", "Fase 6"),
]

for i, (fecha, commit, desc, fase) in enumerate(commits, 4):
    ws4[f'A{i}'] = fecha
    ws4[f'B{i}'] = commit
    ws4[f'C{i}'] = desc
    ws4[f'D{i}'] = fase

ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 55
ws4.column_dimensions['D'].width = 10

# Save
wb.save('Carta_Gantt_Simple_Completo.xlsx')
print("Excel completo generado: Carta_Gantt_Simple_Completo.xlsx")
print("Hojas: Resumen del Proyecto, Carta Gantt, Proximo Sprint, Hitos (Commits)")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Carta Gantt - Simple"

# Colors
fill_done = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
fill_progress = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
fill_pending = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")
fill_header = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
fill_milestone = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
font_white = Font(color="FFFFFF", bold=True)
font_dark = Font(color="37474F", bold=True)

# Title
ws.merge_cells('A1:H1')
ws['A1'] = "Carta Gantt — Proyecto Simple"
ws['A1'].font = Font(size=16, bold=True, color="37474F")
ws.merge_cells('A2:H2')
ws['A2'] = "27 Abril 2026 - 07 Julio 2026"
ws['A2'].font = Font(size=11, color="78909C")

# Headers
headers = ["Fase", "Tarea", "Inicio", "Fin", "Duración", "Estado", "Progreso", "Notas"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.fill = fill_header
    cell.font = font_white
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Phase data
phases = [
    ("Fase 1", "Fundación y Autenticación", "27/04/2026", "28/04/2026", "done", 100, "Login email/Google, reglas Firestore, RoleDispatcher"),
    ("Fase 2", "IA / Súper Experto", "28/04/2026", "29/04/2026", "done", 100, "Cloud Functions Gemini, fallback local, chat UI"),
    ("Fase 3", "Módulo TEA (Pictogramas)", "30/04/2026", "06/05/2026", "done", 100, "Tablero ARASAAC, TTS, pictos custom, pictogramSettings"),
    ("Fase 4", "Módulo TDAH (Tareas y Foco)", "05/05/2026", "09/05/2026", "done", 100, "CRUD tareas, Pomodoro, respiración, sistema dopamina"),
    ("Fase 5", "Integración y Correcciones", "09/05/2026", "23/05/2026", "done", 100, "Vinculación tutor-usuario, historial, roles unificados"),
    ("Fase 6", "Pulido y Testing", "24/05/2026", "16/06/2026", "progress", 75, "Dashboard progreso, control pestañas, fix nav bar"),
    ("Fase 7", "Documentación y Entrega", "17/06/2026", "07/07/2026", "pending", 0, "Manual usuario, APK firmado, video demo, presentación"),
]

for row_idx, phase in enumerate(phases, 5):
    fase, tarea, inicio, fin, estado, progreso, notas = phase
    duracion = (datetime.strptime(fin, "%d/%m/%Y") - datetime.strptime(inicio, "%d/%m/%Y")).days + 1
    
    ws.cell(row=row_idx, column=1, value=fase)
    ws.cell(row=row_idx, column=2, value=tarea)
    ws.cell(row=row_idx, column=3, value=inicio)
    ws.cell(row=row_idx, column=4, value=fin)
    ws.cell(row=row_idx, column=5, value=f"{duracion} días")
    
    estado_cell = ws.cell(row=row_idx, column=6)
    estado_labels = {"done": "✅ Completado", "progress": "🔄 En Curso", "pending": "⏳ Pendiente"}
    estado_cell.value = estado_labels[estado]
    estado_cell.fill = {"done": fill_done, "progress": fill_progress, "pending": fill_pending}[estado]
    estado_cell.font = font_white
    estado_cell.alignment = Alignment(horizontal="center")
    
    ws.cell(row=row_idx, column=7, value=f"{progreso}%")
    ws.cell(row=row_idx, column=8, value=notas)

# Milestones section
ws.cell(row=14, column=1, value="HITOS DEL PROYECTO").font = Font(size=12, bold=True, color="37474F")

milestone_headers = ["Hito", "Fecha", "Descripción", "Estado"]
for col, header in enumerate(milestone_headers, 1):
    cell = ws.cell(row=15, column=col, value=header)
    cell.fill = fill_milestone
    cell.font = font_white
    cell.alignment = Alignment(horizontal="center")

milestones = [
    ("HITO 1", "13/05/2026", "Vinculación Tutor-Usuario por código de 6 caracteres", "done"),
    ("HITO 2", "26/05/2026", "Sincronización Completa tutor ↔ usuario en tiempo real", "done"),
    ("HITO 3", "30/06/2026", "MVP Listo — todas las funcionalidades implementadas", "pending"),
    ("HITO 4", "07/07/2026", "Entrega Final — APK + Documentación + Video Demo", "pending"),
]

for row_idx, m in enumerate(milestones, 16):
    hito, fecha, desc, estado = m
    ws.cell(row=row_idx, column=1, value=hito)
    ws.cell(row=row_idx, column=2, value=fecha)
    ws.cell(row=row_idx, column=3, value=desc)
    
    estado_cell = ws.cell(row=row_idx, column=4)
    estado_labels = {"done": "✅ Alcanzado", "pending": " Pendiente"}
    estado_cell.value = estado_labels[estado]
    estado_cell.fill = fill_done if estado == "done" else fill_pending
    estado_cell.font = font_white
    estado_cell.alignment = Alignment(horizontal="center")

# Column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 55

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

for row in range(4, 12):
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border

for row in range(15, 20):
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border

# Save
wb.save('Carta_Gantt_Simple.xlsx')
print("Excel Gantt chart generated: Carta_Gantt_Simple.xlsx")
